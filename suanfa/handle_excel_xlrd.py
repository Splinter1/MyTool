import xlrd
import os
import openpyxl

path = os.path.dirname(os.path.dirname(__file__))
filepath = path + "/suanfa/"
data_file = filepath + "test_data.xlsx"
result_file = filepath + "result.xlsx"


class HandExcel:
    # 加载excel
    def load_excel(self, file):
        if file == "data":
            open_excel = xlrd.open_workbook(data_file)
            return open_excel
        if file == "result":
            open_excel = xlrd.open_workbook(result_file)
            return open_excel

    def get_sheet_data(self, file, index=None):
        if index is None:
            index = 0
        data = self.load_excel(file).sheet_by_index(index)
        return data

    # 获取一个单元格的数据
    def get_cell_value(self, file, row, cell):
        data = self.get_sheet_data(file).cell(row, cell).value
        return data

    # 获取行数
    def get_rows(self, file):
        rows = self.get_sheet_data(file).nrows
        return rows

    # 获取一行的内容
    def get_rows_value(self, file, row):
        row_list = []
        for i in self.get_sheet_data(file).row(row-1):
            row_list.append(i.value)
        return row_list

    # 获取一列内容
    def get_cols_value(self, file, col=None):
        cols_list = []
        if col is None:
            col = 'B'
        cols_list_data = self.get_sheet_data(file)[col]
        for i in cols_list_data:
            cols_list.append(i.value)
        return cols_list

    # 根据数据搜索列号
    def get_cells_num(self, file, cell_name):
        data_list = self.get_rows_value(file, 1)
        cell_num = data_list.index(cell_name)
        return cell_num

    # 通过数据搜索行号
    def get_rows_num(self, case_id, file):
        num = 1
        cols_data = self.get_cols_value(file)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1

    # 获取excel中所有数据  每行为list
    def get_excel_data(self, file):
        data_list = []
        rows = self.get_rows(file)
        for i in range(rows-1):
            data_list.append(self.get_rows_value(file, i+2))
        return data_list

    # 写入单个数据
    def write_single_data(self, file, row, col, value):
        if file == "data":
            wb = openpyxl.load_workbook(data_file)
        if file == "result":
            wb = openpyxl.load_workbook(result_file)
        wr = wb.active
        wr.cell(row, col, value)
        if file == "bug":
            wb.save(data_file)
        if file == "result":
            wb.save(result_file)

    # 写入一行数据数据，data需要一个二维数组，注意是二维的！
    def write_list_data(self, file, data):
        if file == "data":
            wb = openpyxl.load_workbook(data_file)
        if file == "result":
            wb = openpyxl.load_workbook(result_file)
        wr = wb.active
        for row in data:
            wr.append(row)
        if file == "bug":
            wb.save(data_file)
        if file == "result":
            wb.save(result_file)


if __name__ == '__main__':
    # HandExcel().write_list_data("result", [["0"]])
    # print(HandExcel().get_cell_value("count", 2, 3))
    print(HandExcel().get_cells_num("data", "ficq452b"))
    # print(HandExcel().get_rows('data'))
    # print(HandExcel().get_cols_value())
    # print(HandExcel().get_rows_num("addDevice3"))
    # print(HandExcel().get_excel_data())
    # print(HandExcel().get_cells_num("是否执行"))
    # s = HandExcel().get_cells_num("count", "胡海峰")
    # print(type(s))
    # HandExcel().excel_write_data("count", 2, HandExcel().get_cells_num("count", "胡海峰"), 999)
    # print(HandExcel().get_excel_data("data"))
    # print(csv_file)

